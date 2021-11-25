# code pour importer les qualification afin de les entrer automatiquement dans la BD

import openpyxl
from pathlib import Path

xlsx_file = Path('qualification_LCP.xlsx')
wb_object = openpyxl.load_workbook(xlsx_file)

print("---------------------")
print("wb_object", wb_object)
print("---------------------")

activeSheet = wb_object.active 


print("activeSheet : ", activeSheet)
print("---------------------")

rowNb = activeSheet.max_row
colNb = activeSheet.max_column

print("nombre des lignes = ", rowNb)
print("nombre des colonnes = ", colNb)
print("---------------------")

print("A1 : ", activeSheet["A1"].value)
print("---------------------")

cell = "A" + str(2)
print("cell : ", cell)
print("cell type : ", type(cell))
print("---------------------")


###################################################################
#
# passer cellule par cellule et chercher un mot clefs specifique
#
###################################################################

# create an empty array to store the wanted phrase

array = []

# loop
for row in activeSheet.iter_rows(max_row=rowNb):
    for cell in row:
        cellValue = cell.value
        if(cellValue != None):
            word_index = cellValue.find('Qualification')
            if (word_index != -1):
                # it is a phrase that we want
                # print("cell.value = ", cellValue)
                array.append(cellValue)

            else:
                pass



# Formatting array's phrases to the wanted form
formatted_array = []
for phrases in array:
    #print(phrases[16:-1])
    formatted_phrase = phrases[16:]  # dernuiere lettre manque
    formatted_array.append(formatted_phrase)

# print(formatted_array)

# Formatting array's phrases to inser it into postgreSQL DataBase
postgreSQL_array = []
Id = 241
for postgreSQLPhrase in formatted_array:
    insertLine = "("+ str(Id) + ", '" + postgreSQLPhrase + "'),"
    postgreSQL_array.append(insertLine)
    Id += 1
    print(insertLine)

# Transfer the output to a .txt file
file = open("output.txt", "a", encoding="utf-8")
for phrases in postgreSQL_array:
    file.write(phrases + "\n")
file.close()




